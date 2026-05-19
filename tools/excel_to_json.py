import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_NAME = "البيان المنظم1.xlsx"
FINAL_EXCEL_NAME = "البيان_المنظم1_نهائي_كما_اتفقنا.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "sites.json"
DEFAULT_JS_OUTPUT = PROJECT_ROOT / "data" / "sites.js"


FIELD_ALIASES = {
    "serial": ["م", "رقم", "التسلسل"],
    "region": ["المنطقة"],
    "facilityType": ["نوع المنشأة", "نوع المنشاٴة", "نوع الموقع"],
    "name": ["اسم المنشأة", "اسم المنشاٴة", "اسم الموقع", "المنشأة"],
    "latitude": ["Latitude", "latitude", "lat", "خط العرض"],
    "longitude": ["Longitude", "longitude", "lng", "lon", "خط الطول"],
    "operator": ["الجهة المشغلة", "الجهة", "المشغل"],
    "beds": ["إجمالي الأسرة", "عدد الأسرة", "عدد الاسرة", "الأسرة", "الاسرة", "السعة السريرية"],
    "bedsTotal": ["إجمالي الأسرة", "عدد الأسرة", "عدد الاسرة", "السعة السريرية"],
    "bedsIcu": ["أسرة العناية المركزة", "العناية المركزة"],
    "bedsEmergency": ["أسرة الطوارئ", "الطوارئ"],
    "bedsClinics": ["أسرة العيادات", "العيادات"],
    "bedsOperations": ["أسرة العمليات", "العمليات"],
    "bedsIsolation": ["أسرة العزل", "العزل"],
    "bedsHeat": ["أسرة ضربات الحرارة", "أسرة ضربات الشمس", "ضربات الحرارة", "ضربات الشمس"],
    "morgueCapacity": ["سعة ثلاجة الموتى", "ثلاجة الموتى"],
    "serviceScope": ["مجال الخدمة"],
    "mainServices": ["التخصصات أو الخدمات الرئيسية", "التخصصات والخدمات الرئيسية", "التخصصات الرئيسية"],
    "coverageScope": ["نطاق التغطية"],
    "googleMapsUrl": ["رابط Google Maps", "رابط خرائط قوقل", "رابط الموقع", "خرائط جوجل"],
    "bedsDataType": ["نوع بيانات الأسرة", "نوع بيانات الاسرة"],
    "responsibleName": ["اسم المسؤول", "مسؤول الموقع", "المسؤول"],
    "responsibleMobile": ["رقم جوال المسؤول", "جوال المسؤول", "جوال مسئول الموقع"],
    "centerManagerName": ["اسم مدير/مشرف المركز", "اسم مدير المركز", "مدير المركز"],
    "centerManagerMobile": [
        "جوال مدير/مشرف المركز",
        "رقم جوال مدير المركز",
        "جوال مدير المركز",
    ],
    "custodyReceiverName": ["اسم مستلم العهدة", "مستلم العهدة"],
    "custodyReceiverMobile": ["جوال مستلم العهدة", "رقم جوال مستلم العهدة"],
    "hospitalManagerName": ["مدير المستشفى", "اسم مدير المستشفى"],
    "hospitalManagerMobile": ["جوال مدير المستشفى", "رقم جوال مدير المستشفى"],
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip()
    return "".join(ch for ch in text.casefold() if ch.isalnum())


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def numeric_value(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def find_input_file(explicit_path=None):
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if path.exists():
            return path
        raise FileNotFoundError(f"Excel file not found: {path}")

    candidates = [
        PROJECT_ROOT / FINAL_EXCEL_NAME,
        PROJECT_ROOT / DEFAULT_EXCEL_NAME,
        Path.home() / "Downloads" / DEFAULT_EXCEL_NAME,
    ]
    for path in candidates:
        if path.exists():
            return path

    raise FileNotFoundError(
        f"Could not find {DEFAULT_EXCEL_NAME}. Pass the Excel path as an argument."
    )


def header_index_map(headers):
    normalized_headers = {normalize_header(header): index for index, header in enumerate(headers)}
    mapping = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = normalize_header(alias)
            if key in normalized_headers:
                mapping[field] = normalized_headers[key]
                break
    return mapping


def get_cell(row, mapping, field):
    index = mapping.get(field)
    if index is None or index >= len(row):
        return ""
    return clean_value(row[index])


def read_extra_fields(row, headers, mapping):
    mapped_indexes = set(mapping.values())
    extra = {}
    for index, header in enumerate(headers):
        if index in mapped_indexes:
            continue
        label = clean_value(header)
        if not label or index >= len(row):
            continue
        value = clean_value(row[index])
        if value:
            extra[label] = value
    return extra


def read_sites(excel_path):
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sites = []

    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                continue

            mapping = header_index_map(headers)
            if "latitude" not in mapping or "longitude" not in mapping:
                continue

            for row_number, row in enumerate(rows, start=2):
                latitude = numeric_value(row[mapping["latitude"]])
                longitude = numeric_value(row[mapping["longitude"]])
                if latitude is None or longitude is None:
                    continue

                center_manager = {
                    "name": get_cell(row, mapping, "centerManagerName"),
                    "mobile": get_cell(row, mapping, "centerManagerMobile"),
                }
                responsible = {
                    "name": get_cell(row, mapping, "responsibleName") or center_manager["name"],
                    "mobile": get_cell(row, mapping, "responsibleMobile") or center_manager["mobile"],
                }
                hospital_manager = {
                    "name": get_cell(row, mapping, "hospitalManagerName") or center_manager["name"],
                    "mobile": get_cell(row, mapping, "hospitalManagerMobile") or center_manager["mobile"],
                }

                site = {
                    "id": get_cell(row, mapping, "serial") or f"{sheet.title}-{row_number}",
                    "sourceSheet": sheet.title,
                    "sourceRow": row_number,
                    "name": get_cell(row, mapping, "name"),
                    "region": get_cell(row, mapping, "region"),
                    "facilityType": get_cell(row, mapping, "facilityType"),
                    "operator": get_cell(row, mapping, "operator"),
                    "beds": get_cell(row, mapping, "beds"),
                    "bedsTotal": get_cell(row, mapping, "bedsTotal"),
                    "bedsIcu": get_cell(row, mapping, "bedsIcu"),
                    "bedsEmergency": get_cell(row, mapping, "bedsEmergency"),
                    "bedsClinics": get_cell(row, mapping, "bedsClinics"),
                    "bedsOperations": get_cell(row, mapping, "bedsOperations"),
                    "bedsIsolation": get_cell(row, mapping, "bedsIsolation"),
                    "bedsHeat": get_cell(row, mapping, "bedsHeat"),
                    "morgueCapacity": get_cell(row, mapping, "morgueCapacity"),
                    "serviceScope": get_cell(row, mapping, "serviceScope"),
                    "mainServices": get_cell(row, mapping, "mainServices"),
                    "coverageScope": get_cell(row, mapping, "coverageScope"),
                    "bedsDataType": get_cell(row, mapping, "bedsDataType"),
                    "latitude": latitude,
                    "longitude": longitude,
                    "responsible": responsible,
                    "centerManager": center_manager,
                    "custodyReceiver": {
                        "name": get_cell(row, mapping, "custodyReceiverName"),
                        "mobile": get_cell(row, mapping, "custodyReceiverMobile"),
                    },
                    "hospitalManager": hospital_manager,
                    "extraFields": read_extra_fields(row, headers, mapping),
                    "googleMapsUrl": get_cell(row, mapping, "googleMapsUrl")
                    or f"https://www.google.com/maps?q={latitude},{longitude}",
                }
                sites.append(site)
    finally:
        workbook.close()

    return sites


def convert_excel_to_json(excel_path, output_path=DEFAULT_OUTPUT):
    sites = read_sites(excel_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(sites, ensure_ascii=False, indent=2)
    output_path.write_text(json_text, encoding="utf-8")

    if output_path.resolve() == DEFAULT_OUTPUT.resolve():
        DEFAULT_JS_OUTPUT.write_text(
            "window.SITES_DATA = " + json_text + ";\n",
            encoding="utf-8",
        )
    return sites


def main():
    parser = argparse.ArgumentParser(description="Convert health sites Excel data to JSON.")
    parser.add_argument("excel", nargs="?", help="Path to البيان المنظم1.xlsx")
    parser.add_argument(
        "-o",
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Output JSON path. Default: data/sites.json",
    )
    args = parser.parse_args()

    excel_path = find_input_file(args.excel)
    sites = convert_excel_to_json(excel_path, args.output)
    print(f"Converted {len(sites)} sites to {args.output}")


if __name__ == "__main__":
    main()
